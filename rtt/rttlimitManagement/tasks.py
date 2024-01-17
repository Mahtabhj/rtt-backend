import logging
import traceback
from datetime import datetime
from django.utils import timezone
from django.conf import settings
import re

from django.core.files.storage import default_storage
from django.contrib.auth import get_user_model
import openpyxl
from rtt import celery_app
from rttcore.services.email_service import send_mail_via_mailjet_template

from rttlimitManagement.models import RegulationSubstanceLimit, Exemption, LimitUploadLog, \
    LimitAdditionalAttributeValue, RegulationLimitAttribute
from rttregulation.models.models import RegulatoryFramework, Regulation
from rttsubstance.models import Substance

logger = logging.getLogger(__name__)
User = get_user_model()
CELERY_DEFAULT_QUEUE = settings.CELERY_DEFAULT_QUEUE


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def regulation_substance_limit_upload_task(file_path, uploaded_file_name, limit_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-723
    """
    limit_upload_log = LimitUploadLog.objects.get(id=limit_upload_log_id)
    if limit_upload_log.status != 'in_queue' and limit_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    limit_upload_log.status = 'in_progress'
    limit_upload_log.save()
    try:
        succeed_data_entry_count = 0
        print(f"regulation_substance_limit_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        limit_upload_log.file_url = default_storage.url(file_path)
        limit_upload_log.total_data_in_file = (worksheet_data.max_row * 2) - 2
        limit_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None:
                continue
            try:
                substance_id = int(row[0].value)
                substance = Substance.objects.get(pk=substance_id)
                scope = row[3].value if row[3].value else None
                limit_value = float(row[4].value) if (
                            row[4].value is not None and not isinstance(row[4].value, str)) else None
                measurement_limit_unit = row[5].value if row[5].value else None
                limit_note = row[6].value if row[6].value else None
                date_into_force = row[7].value if row[7].value else None
                try:
                    framework_id = int(row[1].value) if row[1].value else None
                    regulatory_framework = RegulatoryFramework.objects.get(pk=framework_id)
                    obj, created = RegulationSubstanceLimit.objects.get_or_create(
                        substance=substance,
                        regulatory_framework=regulatory_framework,
                        scope=scope, limit_value=limit_value, measurement_limit_unit=measurement_limit_unit,
                        limit_note=limit_note, date_into_force=date_into_force, status='active'
                    )
                    if created:
                        succeed_data_entry_count = succeed_data_entry_count + 1
                except Exception as exc:
                    logger.error(str(exc), exc_info=True)

                try:
                    regulation_id = int(row[2].value) if row[2].value else None
                    regulation = Regulation.objects.get(pk=regulation_id)
                    obj, created = RegulationSubstanceLimit.objects.get_or_create(
                        substance=substance,
                        regulation=regulation,
                        scope=scope, limit_value=limit_value, measurement_limit_unit=measurement_limit_unit,
                        limit_note=limit_note, date_into_force=date_into_force
                    )
                    if created:
                        succeed_data_entry_count = succeed_data_entry_count + 1
                except Exception as exc:
                    logger.error(str(exc), exc_info=True)
            except Exception as exc:
                logger.error(str(exc), exc_info=True)

        limit_upload_log.succeed_data_entry = succeed_data_entry_count
        limit_upload_log.failed_data_entry = (worksheet_data.max_row * 2) - succeed_data_entry_count - 2
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.status = 'success'
        limit_upload_log.save()
        print(f"regulation_substance_limit_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        limit_upload_log.status = 'fail'
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.save()
        raise


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def exemption_upload_task(file_path, uploaded_file_name, email_to, limit_upload_log_id):
    """
    https://chemycal.atlassian.net/browse/RTT-724
    """
    limit_upload_log = LimitUploadLog.objects.get(id=limit_upload_log_id)
    if limit_upload_log.status != 'in_queue' and limit_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    limit_upload_log.status = 'in_progress'
    limit_upload_log.save()
    try:
        succeed_data_entry_count = 0
        print(f"exemption_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        limit_upload_log.file_url = default_storage.url(file_path)
        limit_upload_log.total_data_in_file = (worksheet_data.max_row * 2) - 2
        limit_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[2].value is None:
                continue
            try:
                substance_id = int(row[2].value)
                substance = Substance.objects.get(pk=substance_id)
                article = row[3].value if row[3].value else None
                reference = row[4].value if row[4].value else None
                application = row[5].value if row[5].value else None
                expiration_date = row[6].value if row[6].value else None
                date_into_force = row[7].value if row[7].value else None
                notes = row[8].value if row[8].value else None
                try:
                    framework_id = int(row[0].value) if row[0].value else None
                    regulatory_framework = RegulatoryFramework.objects.get(pk=framework_id)
                    obj, created = Exemption.objects.get_or_create(
                        substance=substance,
                        regulatory_framework=regulatory_framework,
                        article=article, reference=reference, application=application, expiration_date=expiration_date,
                        date_into_force=date_into_force, notes=notes
                    )
                    if created:
                        succeed_data_entry_count = succeed_data_entry_count + 1
                except Exception as exc:
                    logger.error(str(exc), exc_info=True)

                try:
                    regulation_id = int(row[1].value) if row[1].value else None
                    regulation = Regulation.objects.get(pk=regulation_id)
                    obj, created = Exemption.objects.get_or_create(
                        substance=substance,
                        regulation=regulation,
                        article=article, reference=reference, application=application, expiration_date=expiration_date,
                        date_into_force=date_into_force, notes=notes
                    )
                    if created:
                        succeed_data_entry_count = succeed_data_entry_count + 1
                except Exception as exc:
                    logger.error(str(exc), exc_info=True)
                # send Exemption add report via email to the user
                subject = f'Exemption add report'
                variables_dict = {
                    'date': str(timezone.now().date()),
                    'upload_report': f"{succeed_data_entry_count} exemption(s) have been added successfully",
                    'file_name': uploaded_file_name
                }
                template_id = settings.MAILJET_EXEMPTION_ADD_REPORT_TEMPLATE_ID
                send_mail_via_mailjet_template(email_to, template_id, subject, variables_dict)
            except Exception as exc:
                logger.error(str(exc), exc_info=True)

        limit_upload_log.succeed_data_entry = succeed_data_entry_count
        limit_upload_log.failed_data_entry = (worksheet_data.max_row * 2) - succeed_data_entry_count - 2
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.status = 'success'
        limit_upload_log.save()
        print(f"exemption_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        limit_upload_log.status = 'fail'
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.save()
        raise


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def limit_additional_attribute_value_upload_task(file_path, uploaded_file_name, limit_upload_log_id):
    """
    https://chemycal.atlassian.net/browse/RTT-722
    """
    limit_upload_log = LimitUploadLog.objects.get(id=limit_upload_log_id)
    if limit_upload_log.status != 'in_queue' and limit_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    limit_upload_log.status = 'in_progress'
    limit_upload_log.save()
    try:
        succeed_data_entry_count = 0
        print(f"limit_additional_attribute_value_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        limit_upload_log.file_url = default_storage.url(file_path)
        limit_upload_log.total_data_in_file = worksheet_data.max_row - 1
        limit_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[1].value is None:
                continue
            try:
                regulation_substance_limit_id = int(row[0].value)
                regulation_limit_attribute_id = int(row[1].value)
                value = row[2].value
                regulation_substance_limit = RegulationSubstanceLimit.objects.get(pk=regulation_substance_limit_id)
                regulation_limit_attribute = RegulationLimitAttribute.objects.get(pk=regulation_limit_attribute_id)
                obj, created = LimitAdditionalAttributeValue.objects.update_or_create(
                    regulation_substance_limit=regulation_substance_limit,
                    regulation_limit_attribute=regulation_limit_attribute,
                    defaults={'value': value},
                )
                succeed_data_entry_count += 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        limit_upload_log.succeed_data_entry = succeed_data_entry_count
        limit_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.status = 'success'
        limit_upload_log.save()
        print(f"limit_additional_attribute_value_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        limit_upload_log.status = 'fail'
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.save()
        raise


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def limit_with_additional_attribute_value_upload_task(user_id, file_path, uploaded_file_name, limit_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-940
    """
    limit_upload_log = LimitUploadLog.objects.get(id=limit_upload_log_id)
    if limit_upload_log.status != 'in_queue' and limit_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    limit_upload_log.status = 'in_progress'
    limit_upload_log.save()
    try:
        succeed_data_entry_count = 0
        print(f"limit_with_additional_attribute_value_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        limit_upload_log.file_url = default_storage.url(file_path)
        limit_upload_log.total_data_in_file = worksheet_data.max_row - 2
        limit_upload_log.save()
        count = 0
        failed_substances = []
        has_fail = False
        pattern = re.compile("(^[a-zA-Z :]{0,}[0-9]{1,})(-[0-9]{1,})(-[0-9]{1,})$")
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or count == 2 or (row[0].value is None and row[1].value is None):
                continue
            try:
                substance = None
                if row[0].value:
                    substance = Substance.objects.filter(id=int(row[0].value)).first()
                elif not substance and row[1].value:
                    substance = Substance.objects.filter(cas_no__iexact=str(row[1].value)).first()
                if not substance:
                    fail_message = ''
                    if row[0].value:
                        fail_message = 'No substance found with this family_id'
                    if row[1].value:
                        is_valid_cas = bool(pattern.match(str(row[1].value)))
                        if not is_valid_cas:
                            fail_message += ', ' if len(fail_message) > 0 else ''
                            fail_message += 'Invalid CAS'
                        else:
                            fail_message += ', ' if len(fail_message) > 0 else ''
                            fail_message += 'No substance found with this CAS'
                    if len(fail_message) > 0:
                        has_fail = True
                        failed_substances.append({
                            'family_id': str(row[0].value) if row[0].value else '',
                            'cas_no': str(row[1].value) if row[1].value else '',
                            'reason': fail_message
                        })
                if row[2].value:
                    framework = RegulatoryFramework.objects.get(id=int(row[2].value))
                    regulation_substance_limit, created = RegulationSubstanceLimit.objects.get_or_create(
                        substance_id=substance.id,
                        regulatory_framework=framework,
                        scope=row[4].value,
                        limit_value=float(row[5].value) if row[5].value else None,
                        measurement_limit_unit=row[6].value,
                        limit_note=row[7].value,
                        date_into_force=row[8].value,
                        status='active', defaults={'modified': row[9].value}
                    )
                elif row[3].value:
                    regulation = Regulation.objects.get(id=int(row[3].value))
                    regulation_substance_limit, created = RegulationSubstanceLimit.objects.get_or_create(
                        substance_id=substance.id,
                        regulation=regulation,
                        scope=row[4].value,
                        limit_value=float(row[5].value) if row[5].value else None,
                        measurement_limit_unit=row[6].value,
                        limit_note=row[7].value,
                        date_into_force=row[8].value,
                        status='active', defaults={'modified': row[9].value}
                    )
                else:
                    continue
                if created:
                    succeed_data_entry_count += 1
                for idx in range(10, 20, 2):
                    try:
                        value_idx = idx + 1
                        reg_limit_att_id = int(row[idx].value)
                        regulation_limit_attribute = RegulationLimitAttribute.objects.get(pk=reg_limit_att_id)
                        LimitAdditionalAttributeValue.objects.update_or_create(
                            regulation_substance_limit=regulation_substance_limit,
                            regulation_limit_attribute=regulation_limit_attribute,
                            defaults={'value': row[value_idx].value},
                        )
                    except Exception as exc:
                        logger.error(str(exc), exc_info=True)
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        if failed_substances:
            limit_upload_log.traceback = failed_substances
        limit_upload_log.succeed_data_entry = succeed_data_entry_count
        limit_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 2
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.status = 'success'
        limit_upload_log.save()
        # send substance_add report to the user
        email_to = User.objects.filter(id=user_id).first().email
        subject = f'Limit with Additional Attribute Value add report'
        variables_dict = {
            'date': str(timezone.now().date()),
            'upload_report': f"{succeed_data_entry_count} limit(s) have been added",
            'file_name': uploaded_file_name,
            'has_fail': has_fail,
            'failed_substances': failed_substances
        }
        template_id = settings.MAILJET_LIMIT_WITH_ADDITIONAL_ATTRIBUTE_VALUE_UPLOAD_TEMPLATE_ID
        send_mail_via_mailjet_template(email_to, template_id, subject, variables_dict)
        print(f"limit_with_additional_attribute_value_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        limit_upload_log.status = 'fail'
        limit_upload_log.end_time = datetime.now()
        limit_upload_log.traceback = str(traceback.format_exc())
        limit_upload_log.save()
        raise
