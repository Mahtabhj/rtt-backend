from hashlib import sha256
import logging
from datetime import datetime
import requests
import traceback
from django.utils import timezone
import re

from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.contrib.auth import get_user_model
from django.conf import settings
import openpyxl
from rtt import celery_app
from rttcore.services.email_service import send_mail_via_mailjet_template
from rttsubstance.models import Substance, SubstanceUploadLog, SubstanceUsesAndApplication, \
    SubstancePropertyDataPoint, PropertyDataPoint, SubstanceFamily, UserSubstanceAddLog
from rttregulation.models.models import RegulatoryFramework, Regulation, SubstanceRegulatoryFramework, \
    SubstanceRegulation, RegulationMilestone
from rttproduct.models.models import Product

logger = logging.getLogger(__name__)
User = get_user_model()
CELERY_DEFAULT_QUEUE = settings.CELERY_DEFAULT_QUEUE


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def substance_basic_details_upload_task(file_path, uploaded_file_name, substance_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-683
    """

    substance_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_upload_log.status != 'in_queue' and substance_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_upload_log.status = 'in_progress'
    substance_upload_log.save()

    try:
        succeed_data_entry_count = 0
        print(f"substance_basic_details_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_upload_log.file_url = default_storage.url(file_path)
        substance_upload_log.total_data_in_file = worksheet_data.max_row - 1
        substance_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[4].value is None:
                continue
            try:
                Substance.objects.update_or_create(
                    chemycal_id__iexact=row[4].value,
                    defaults={'name': row[0].value,
                              'ec_no': row[1].value if row[1].value else None,
                              'cas_no': row[2].value if row[2].value else None,
                              'chemycal_id': row[4].value,
                              'molecular_formula': row[3].value if row[3].value else None},
                )
                succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        substance_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.status = 'success'
        substance_upload_log.save()
        print(f"substance_basic_details_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_upload_log.status = 'fail'
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def substances_uses_and_applications_upload_task(file_path, uploaded_file_name, substance_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-684
    """

    substance_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_upload_log.status != 'in_queue' and substance_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_upload_log.status = 'in_progress'
    substance_upload_log.save()

    try:
        succeed_data_entry_count = 0
        print(f"substances_uses_and_applications_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_upload_log.file_url = default_storage.url(file_path)
        substance_upload_log.total_data_in_file = worksheet_data.max_row - 1
        substance_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[1].value is None:
                continue
            try:
                substance_id = int(row[0].value) if row[0].value else None
                use_and_app_id = int(row[1].value) if row[1].value else None
                if not substance_id or not use_and_app_id:
                    continue
                substance = Substance.objects.get(pk=substance_id)
                use_and_app = SubstanceUsesAndApplication.objects.get(pk=use_and_app_id)
                use_and_app.substances.add(substance)
                succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        substance_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.status = 'success'
        substance_upload_log.save()
        print(f"substances_uses_and_applications_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_upload_log.status = 'fail'
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def substances_frameworks_or_regulations_upload_task(file_path, uploaded_file_name, substance_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-687
    """
    substance_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_upload_log.status != 'in_queue' and substance_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_upload_log.status = 'in_progress'
    substance_upload_log.save()

    try:
        succeed_data_entry_count = 0
        print(f"substances_frameworks_or_regulations_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_upload_log.file_url = default_storage.url(file_path)
        substance_upload_log.total_data_in_file = (worksheet_data.max_row * 2) - 2
        substance_upload_log.save()
        count = 0
        substance_regulatory_framework_list = []
        substance_regulation_list = []
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[2].value is None:
                continue
            substance_id = int(row[2].value)
            try:
                substance = Substance.objects.get(pk=substance_id)
                framework_id = int(row[0].value) if row[0].value else None
                regulatory_framework = RegulatoryFramework.objects.get(pk=framework_id)
                regulatory_framework.substances.add(substance)
                succeed_data_entry_count = succeed_data_entry_count + 1
                '''
                For Substance RegulatoryFramework relation added date
                '''
                substance_regulatory_framework_list.append(
                    SubstanceRegulatoryFramework(regulatory_framework_id=framework_id,
                                                 substance_id=substance_id))
            except Exception as exc:
                logger.error(str(exc), exc_info=True)

            try:
                substance = Substance.objects.get(pk=substance_id)
                regulation_id = int(row[1].value) if row[1].value else None
                regulation = Regulation.objects.get(pk=regulation_id)
                regulation.substances.add(substance)
                succeed_data_entry_count = succeed_data_entry_count + 1

                '''
                For Substance Regulation relation added date
                '''
                substance_regulation_list.append(SubstanceRegulation(regulation_id=regulation_id,
                                                                     substance_id=substance_id))
            except Exception as exc:
                logger.error(str(exc), exc_info=True)

        SubstanceRegulatoryFramework.objects.bulk_create(substance_regulatory_framework_list, ignore_conflicts=True)
        SubstanceRegulation.objects.bulk_create(substance_regulation_list, ignore_conflicts=True)

        substance_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_upload_log.failed_data_entry = (worksheet_data.max_row * 2) - succeed_data_entry_count - 2
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.status = 'success'
        substance_upload_log.save()
        print(f"substances_frameworks_or_regulations_upload_task ..........success..........")
    except Exception as ex:
        logger.error(str(ex), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_upload_log.status = 'fail'
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def substance_related_products_upload_task(file_path, uploaded_file_name, substance_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-685
    """
    substance_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_upload_log.status != 'in_queue' and substance_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_upload_log.status = 'in_progress'
    substance_upload_log.save()

    try:
        succeed_data_entry_count = 0
        print(f"substance_related_products_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_upload_log.total_data_in_file = worksheet_data.max_row - 1
        substance_upload_log.file_url = default_storage.url(file_path)
        substance_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[1].value is None:
                continue
            try:
                substance_id = int(row[0].value) if row[0].value else None
                product_id = int(row[1].value) if row[1].value else None
                if not substance_id or not product_id:
                    continue
                substance = Substance.objects.get(pk=substance_id)
                product = Product.objects.get(pk=product_id)
                product.substances.add(substance)
                succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        substance_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.status = 'success'
        substance_upload_log.save()
        print(f"substance_related_products_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_upload_log.status = 'fail'
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def substance_data_upload_task(file_path, uploaded_file_name, substance_upload_log_id, existing_data_col=False):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-686
    doc: https://chemycal.atlassian.net/browse/RTT-1427
    """
    substance_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_upload_log.status != 'in_queue' and substance_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_upload_log.status = 'in_progress'
    substance_upload_log.save()
    try:
        succeed_data_entry_count = 0
        print(f"substance_data_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_upload_log.total_data_in_file = worksheet_data.max_row - 1
        substance_upload_log.file_url = default_storage.url(file_path)
        substance_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[1].value is None or row[2].value is None:
                continue
            try:
                substance_id = int(row[0].value) if row[0].value else None
                property_data_point_id = int(row[1].value) if row[1].value else None
                value = str(row[2].value) if row[2].value else None
                image = None
                try:
                    response = requests.get(str(row[3].value), allow_redirects=False)
                    seed_str = response.url + str(datetime.now())
                    unique_name = sha256(seed_str.encode()).hexdigest()
                    unique_name = unique_name[:20]
                    image_path = f'media/substance_property_data_point/{unique_name}.png'
                    image = default_storage.save(image_path, ContentFile(response.content))
                except Exception as exc:
                    logger.error(str(exc), exc_info=True)
                    image = None
                substance = Substance.objects.get(pk=substance_id)
                modified_date = row[4].value if row[4].value else datetime.now()
                property_data_point = PropertyDataPoint.objects.get(pk=property_data_point_id)
                if existing_data_col:
                    existing_data_decision = row[5].value if row[5].value else None
                    if existing_data_decision.lower() == 'o':
                        # O for overwrite
                        obj, created = SubstancePropertyDataPoint.objects.update_or_create(
                            substance=substance, property_data_point=property_data_point,
                            defaults={'value': value, 'image': image, 'modified': modified_date},
                        )

                    elif existing_data_decision.lower() == 'n':
                        # N for new version
                        try:
                            existing_data = SubstancePropertyDataPoint.objects.get(
                                substance=substance, property_data_point=property_data_point, status='active')
                            existing_data.status = 'deleted'
                            existing_data.save()
                        except Exception as e:
                            logger.error(str(e), exc_info=True)
                        SubstancePropertyDataPoint.objects.create(
                            substance=substance, property_data_point=property_data_point,
                            value=value,image=image, modified=modified_date)
                    else:
                        # skip the record
                        pass
                else:
                    obj, created = SubstancePropertyDataPoint.objects.update_or_create(
                        substance=substance, property_data_point=property_data_point,
                        defaults={'value': value, 'image': image, 'modified': modified_date},
                    )
                succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        substance_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.status = 'success'
        substance_upload_log.save()
        print(f"substance_data_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_upload_log.status = 'fail'
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def substance_families_upload_task(file_path, uploaded_file_name, substance_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-697
    """
    substance_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_upload_log.status != 'in_queue' and substance_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_upload_log.status = 'in_progress'
    substance_upload_log.save()

    try:
        succeed_data_entry_count = 0
        print(f"substance_families_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_upload_log.total_data_in_file = worksheet_data.max_row - 1
        substance_upload_log.file_url = default_storage.url(file_path)
        substance_upload_log.save()
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[1].value is None:
                continue
            try:
                substance_id = int(row[0].value) if row[0].value else None
                family_id = int(row[1].value) if row[1].value else None
                family_source = str(row[2].value) if row[2].value else None
                if not substance_id or not family_id:
                    continue
                if Substance.objects.filter(id=family_id, is_family=True).exists():
                    obj, create = SubstanceFamily.objects.update_or_create(
                        substance_id=substance_id, family_id=family_id,
                        defaults={'family_source': family_source},
                    )
                    succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        substance_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.status = 'success'
        substance_upload_log.save()
        print(f"substance_related_products_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_upload_log.status = 'fail'
        substance_upload_log.end_time = datetime.now()
        substance_upload_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def user_substances_add_task(user_id, organization_id, file_path, uploaded_file_name, uses_and_applications,
                             user_substance_add_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-836
    """
    user_substance_add_log = UserSubstanceAddLog.objects.get(id=user_substance_add_log_id)
    try:
        print(f"substances_add_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        user_substance_add_log.file_url = default_storage.url(file_path)
        user_substance_add_log.save()
        count = 0
        use_and_app_count = 0
        use_and_application_ids = [int(use_and_app_id) for use_and_app_id in uses_and_applications.split(',')]
        valid_substance_ids = []
        failed_substances = []
        has_fail = False
        pattern = re.compile("(^[a-zA-Z :]{0,}[0-9]{1,})(-[0-9]{1,})(-[0-9]{1,})$")
        is_valid_cas = True
        is_valid_ec = True
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or (row[0].value is None and row[1].value is None):
                continue
            fail_message = None
            try:
                ec_no = str(row[0].value) if row[0].value else None
                cas_no = str(row[1].value) if row[1].value else None
                substance_qs = []
                if cas_no:
                    substance_qs = list(Substance.objects.filter(cas_no__iexact=cas_no).values_list('id', flat=True))
                if len(substance_qs) == 0 and ec_no:
                    substance_qs = list(Substance.objects.filter(ec_no__iexact=ec_no).values_list('id', flat=True))
                if len(substance_qs) > 0:
                    valid_substance_ids.extend(substance_qs)
                else:
                    has_fail = True
                    if cas_no:
                        is_valid_cas = bool(pattern.match(cas_no))
                        if not is_valid_cas:
                            fail_message = 'Invalid CAS'
                        else:
                            fail_message = 'No substance found with this CAS'
                    if ec_no:
                        is_valid_ec = bool(pattern.match(ec_no))
                        if not is_valid_ec:
                            if not cas_no:
                                fail_message = 'Invalid EC'
                            elif not is_valid_cas:
                                fail_message += ' and EC'
                            else:
                                fail_message += ', Invalid EC'
                        else:
                            if not cas_no:
                                fail_message = 'No substance found with this EC'
                            elif not is_valid_cas:
                                fail_message += ', No substance found with this EC'
                            else:
                                fail_message += ' and EC'
                    failed_substances.append({
                        'cas_no': cas_no,
                        'ec_no': ec_no,
                        'reason': fail_message
                    })
            except Exception as ex:
                logger.error(str(ex), exc_info=True)
        for use_and_app in use_and_application_ids:
            try:
                use_and_app_qs = SubstanceUsesAndApplication.objects.get(id=use_and_app,
                                                                         organization_id=organization_id)
                use_and_app_qs.substances.add(*valid_substance_ids)
                use_and_app_count += 1
            except Exception as e:
                logger.error(str(e), exc_info=True)
        if failed_substances:
            user_substance_add_log.traceback = failed_substances
        user_substance_add_log.object_ids = uses_and_applications
        user_substance_add_log.substance_count = len(set(valid_substance_ids))
        user_substance_add_log.status = 'success'
        user_substance_add_log.save()
        # send substance_add report to the user
        email_to = User.objects.filter(id=user_id).first().email
        subject = f'Substance add report'
        variables_dict = {
            'date': str(timezone.now().date()),
            'upload_report': f"{user_substance_add_log.substance_count} substance(s) has been added in {use_and_app_count} "
                             f"uses and applications",
            'file_name': uploaded_file_name,
            'has_fail': has_fail,
            'failed_substances': failed_substances
        }
        template_id = settings.MAILJET_SUBSTANCE_ADD_REPORT_TEMPLATE_ID
        send_mail_via_mailjet_template(email_to, template_id, subject, variables_dict)
        print(f"substances_add_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        user_substance_add_log.status = 'fail'
        user_substance_add_log.traceback = str(traceback.format_exc())
        user_substance_add_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def upload_substances_add_relation_task(file_path, uploaded_file_name, process_type, object_id, email_to):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-876
    """
    # TODO: add entry in UserSubstanceAddLog
    try:
        print(f"substances_add_relation_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        count = 0
        valid_substance_ids = []
        failed_substances = []
        has_fail = False
        pattern = re.compile("(^[a-zA-Z :]{0,}[0-9]{1,})(-[0-9]{1,})(-[0-9]{1,})$")
        is_valid_cas = True
        is_valid_ec = True
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or (row[0].value is None and row[1].value is None):
                continue
            fail_message = None
            try:
                ec_no = str(row[0].value) if row[0].value else None
                cas_no = str(row[1].value) if row[1].value else None
                substance_qs = []
                if cas_no:
                    substance_qs = list(Substance.objects.filter(cas_no__iexact=cas_no).values_list('id', flat=True))
                if len(substance_qs) == 0 and ec_no:
                    substance_qs = list(Substance.objects.filter(ec_no__iexact=ec_no).values_list('id', flat=True))
                if len(substance_qs) > 0:
                    valid_substance_ids.extend(substance_qs)
                else:
                    has_fail = True
                    if cas_no:
                        is_valid_cas = bool(pattern.match(cas_no))
                        if not is_valid_cas:
                            fail_message = 'Invalid CAS'
                        else:
                            fail_message = 'No substance found with this CAS'
                    if ec_no:
                        is_valid_ec = bool(pattern.match(ec_no))
                        if not is_valid_ec:
                            if not cas_no:
                                fail_message = 'Invalid EC'
                            elif not is_valid_cas:
                                fail_message += ' and EC'
                            else:
                                fail_message += ', Invalid EC'
                        else:
                            if not cas_no:
                                fail_message = 'No substance found with this EC'
                            elif not is_valid_cas:
                                fail_message += ', No substance found with this EC'
                            else:
                                fail_message += ' and EC'
                    failed_substances.append({
                        'cas_no': cas_no,
                        'ec_no': ec_no,
                        'reason': fail_message
                    })
            except Exception as ex:
                logger.error(str(ex), exc_info=True)
        add_status = 'failed'
        try:
            if process_type == 'regulation':
                regulation_queryset = Regulation.objects.get(id=object_id)
                regulation_queryset.substances.add(*valid_substance_ids)
                add_status = 'succeeded'
            elif process_type == 'regulatory_framework':
                regulatory_framework_qs = RegulatoryFramework.objects.get(id=object_id)
                regulatory_framework_qs.substances.add(*valid_substance_ids)
                add_status = 'succeeded'
            elif process_type == 'milestone':
                milestone_qs = RegulationMilestone.objects.get(id=object_id)
                milestone_qs.substances.add(*valid_substance_ids)
                add_status = 'succeeded'
        except Exception as e:
            logger.error(str(e), exc_info=True)
        # send substance_add report to the user
        subject = f'Report for substance add in {process_type}'
        variables_dict = {
            'date': str(timezone.now().date()),
            'upload_report': f"{len(set(valid_substance_ids))} substance(s) add in {process_type} {add_status}",
            'file_name': uploaded_file_name,
            'has_fail': has_fail,
            'failed_substances': failed_substances
        }
        template_id = settings.MAILJET_SUBSTANCE_ADD_REPORT_TEMPLATE_ID
        send_mail_via_mailjet_template(email_to, template_id, subject, variables_dict)
        print(f"substances_add_relation_task ..........success..........")
    except Exception as exc:
        logger.error(str(exc), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def admin_substance_data_upload_task(user_id, file_path, uploaded_file_name, substance_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-1174
    """
    substance_data_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_data_upload_log.status != 'in_queue' and substance_data_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_data_upload_log.status = 'in_progress'
    substance_data_upload_log.save()
    try:
        succeed_data_entry_count = 0
        print(f"substance_data_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_data_upload_log.total_data_in_file = worksheet_data.max_row - 2
        substance_data_upload_log.file_url = default_storage.url(file_path)
        substance_data_upload_log.save()
        count = 0
        failed_substance_data_list = []
        has_fail = False
        pattern = re.compile("(^[a-zA-Z :]{0,}[0-9]{1,})(-[0-9]{1,})(-[0-9]{1,})$")
        for row in worksheet_data.iter_rows():
            count += 1
            if count < 3 or (row[0].value is None and row[1].value is None) or row[2].value is None or \
                    row[3].value is None:
                continue
            try:
                substance = None
                if row[0].value:
                    substance = Substance.objects.filter(id=int(row[0].value)).first()
                if not substance and row[1].value:
                    substance = Substance.objects.filter(cas_no__iexact=row[1].value).first()

                property_data_point = None
                property_data_point_id = int(row[2].value)
                property_data_point = PropertyDataPoint.objects.filter(id=property_data_point_id).first()

                fail_message = ''
                if not substance:
                    if row[0].value:
                        fail_message = 'No substance found with this family_id'
                    if row[1].value:
                        is_valid_cas = bool(pattern.match(str(row[1].value)))
                        if not is_valid_cas:
                            fail_message += ', ' if len(fail_message) > 0 else ''
                            fail_message += 'Invalid CAS'
                        else:
                            fail_message += ', ' if len(fail_message) > 0 else ''
                            fail_message += 'No substance found with this CAS'
                if not property_data_point:
                    fail_message += ', ' if len(fail_message) > 0 else ''
                    fail_message += 'No Property Data Point found with this data_point_id'
                if len(fail_message) > 0:
                    has_fail = True
                    failed_substance_data_list.append({
                        'family_id': str(row[0].value) if row[0].value else '',
                        'cas_no': str(row[1].value) if row[1].value else '',
                        'data_point_id': str(row[2].value),
                        'reason': fail_message
                    })

                value = str(row[3].value) if row[3].value else None
                image = None
                try:
                    response = requests.get(str(row[4].value), allow_redirects=False)
                    seed_str = response.url + str(datetime.now())
                    unique_name = sha256(seed_str.encode()).hexdigest()
                    unique_name = unique_name[:20]
                    image_path = f'media/substance_property_data_point/{unique_name}.png'
                    image = default_storage.save(image_path, ContentFile(response.content))
                except Exception as exc:
                    logger.error(str(exc), exc_info=True)
                    image = None

                modified_date = row[5].value if row[5].value else datetime.now()
                existing_data_decision = row[6].value if row[6].value else None
                if existing_data_decision.lower() == 'o':
                    # O for overwrite
                    obj, created = SubstancePropertyDataPoint.objects.update_or_create(
                        substance=substance, property_data_point=property_data_point, status='active',
                        defaults={'value': value, 'image': image, 'modified': modified_date},
                    )

                elif existing_data_decision.lower() == 'n':
                    # N for new version
                    try:
                        existing_data = SubstancePropertyDataPoint.objects.get(
                            substance=substance, property_data_point=property_data_point, status='active')
                        existing_data.status = 'deleted'
                        existing_data.save()
                    except Exception as e:
                        logger.error(str(e), exc_info=True)
                    SubstancePropertyDataPoint.objects.create(
                        substance=substance, property_data_point=property_data_point,
                        value=value, image=image, modified=modified_date)
                else:
                    # skip the record
                    pass
                succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        substance_data_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_data_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 2
        substance_data_upload_log.end_time = datetime.now()
        substance_data_upload_log.status = 'success'
        substance_data_upload_log.save()
        # send substance_data_upload report to the user
        email_to = User.objects.filter(id=user_id).first().email
        subject = f'Substance data upload report'
        variables_dict = {
            'date': str(timezone.now().date()),
            'upload_report': f"{succeed_data_entry_count} substance data(s) have been uploaded",
            'file_name': uploaded_file_name,
            'has_fail': has_fail,
            'failed_substance_data_list': failed_substance_data_list
        }
        template_id = settings.MAILJET_SUBSTANCE_DATA_UPLOAD_REPORT_TEMPLATE_ID
        send_mail_via_mailjet_template(email_to, template_id, subject, variables_dict)
        print(f"substance_data_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_data_upload_log.status = 'fail'
        substance_data_upload_log.end_time = datetime.now()
        substance_data_upload_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def user_product_related_substances_upload_task(user_id, file_path, uploaded_file_name, product_id,
                                                user_substance_add_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-1247
    """
    user_substance_add_log = UserSubstanceAddLog.objects.get(id=user_substance_add_log_id)

    try:
        succeed_data_entry_count = 0
        print(f"product_related_substances_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        user_substance_add_log.total_data_in_file = worksheet_data.max_row - 1
        user_substance_add_log.file_url = default_storage.url(file_path)
        user_substance_add_log.status = 'in_progress'
        user_substance_add_log.save()
        count = 0
        product = Product.objects.get(pk=product_id)
        failed_substances = []
        has_fail = False
        fail_message = None
        pattern = re.compile("(^[a-zA-Z :]{0,}[0-9]{1,})(-[0-9]{1,})(-[0-9]{1,})$")
        is_valid_cas = True
        is_valid_ec = True

        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or (row[0].value is None and row[1].value is None):
                continue
            try:
                substance = None
                ec_no = row[0].value if row[0].value else None
                cas_no = row[1].value if row[1].value else None

                if cas_no:
                    substance = Substance.objects.filter(cas_no=cas_no).first()
                if not substance and ec_no:
                    substance = Substance.objects.filter(ec_no=ec_no).first()
                if not substance:
                    has_fail = True
                    if cas_no:
                        is_valid_cas = bool(pattern.match(cas_no))
                        if not is_valid_cas:
                            fail_message = 'Invalid CAS'
                        else:
                            fail_message = 'No substance found with this CAS'
                    if ec_no:
                        is_valid_ec = bool(pattern.match(ec_no))
                        if not is_valid_ec:
                            if not cas_no:
                                fail_message = 'Invalid EC'
                            elif not is_valid_cas:
                                fail_message += ' and EC'
                            else:
                                fail_message += ', Invalid EC'
                        else:
                            if not cas_no:
                                fail_message = 'No substance found with this EC'
                            elif not is_valid_cas:
                                fail_message += ', No substance found with this EC'
                            else:
                                fail_message += ' and EC'
                    failed_substances.append({
                        'cas_no': cas_no,
                        'ec_no': ec_no,
                        'reason': fail_message
                    })
                    continue
                product.substances.add(substance)
                succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)
        if failed_substances:
            user_substance_add_log.traceback = failed_substances

        user_substance_add_log.succeed_data_entry = succeed_data_entry_count
        user_substance_add_log.substance_count = succeed_data_entry_count
        user_substance_add_log.object_ids = product_id
        user_substance_add_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        user_substance_add_log.status = 'success'
        user_substance_add_log.save()

        # send user_substance_add report to the user
        email_to = User.objects.filter(id=user_id).first().email
        subject = f'Substance add report'
        variables_dict = {
            'date': str(timezone.now().date()),
            'upload_report': f"{user_substance_add_log.succeed_data_entry} substance(s) has been added to product: "
                             f"{product.name}.",
            'file_name': uploaded_file_name,
            'has_fail': has_fail,
            'failed_substances': failed_substances
        }
        template_id = settings.MAILJET_PRODUCTS_RELATED_SUBSTANCE_ADD_REPORT_TEMPLATE_ID
        send_mail_via_mailjet_template(email_to, template_id, subject, variables_dict)
        print(f"product_related_substances_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        user_substance_add_log.status = 'fail'
        user_substance_add_log.save()


@celery_app.task(queue=CELERY_DEFAULT_QUEUE)
def admin_substance_families_upload_task(user_id, file_path, uploaded_file_name, family_id, substance_upload_log_id):
    """
    doc: https://chemycal.atlassian.net/browse/RTT-1214
    """
    substance_family_upload_log = SubstanceUploadLog.objects.get(id=substance_upload_log_id)
    if substance_family_upload_log.status != 'in_queue' and substance_family_upload_log.status != 'fail':
        return {'message': 'This task has executed before'}
    substance_family_upload_log.status = 'in_progress'
    substance_family_upload_log.save()
    try:
        succeed_data_entry_count = 0
        print(f"substance_family_upload_task ..........start..........")
        file = default_storage.open(file_path)
        worksheet_data = openpyxl.load_workbook(file).worksheets[0]
        substance_family_upload_log.total_data_in_file = worksheet_data.max_row - 1
        substance_family_upload_log.file_url = default_storage.url(file_path)
        substance_family_upload_log.save()
        count = 0
        failed_substance_family_list = []
        substance_family = Substance.objects.filter(id=family_id).first()
        has_fail = False
        pattern = re.compile("(^[a-zA-Z :]{0,}[0-9]{1,})(-[0-9]{1,})(-[0-9]{1,})$")
        for row in worksheet_data.iter_rows():
            count += 1
            if count < 2:
                continue
            try:
                substance = None
                ec_no = row[0].value if row[0].value else None
                cas_no = row[1].value if row[1].value else None

                if cas_no:
                    substance = Substance.objects.filter(cas_no__iexact=cas_no).first()
                if not substance and ec_no:
                    substance = Substance.objects.filter(ec_no__iexact=ec_no).first()

                fail_message = ''
                if not substance and cas_no:
                    has_fail = True
                    is_valid_cas = bool(pattern.match(cas_no))
                    if not is_valid_cas:
                        fail_message = 'Invalid CAS'
                    else:
                        fail_message = 'No substance found with this CAS'

                if len(fail_message) > 0:
                    failed_substance_family_list.append({
                        'cas_no': row[1].value if row[1].value else '',
                        'reason': fail_message
                    })
                SubstanceFamily.objects.update_or_create(substance=substance, family=substance_family)
                succeed_data_entry_count = succeed_data_entry_count + 1
            except Exception as ex:
                logger.error(str(ex), exc_info=True)

        substance_family_upload_log.succeed_data_entry = succeed_data_entry_count
        substance_family_upload_log.failed_data_entry = worksheet_data.max_row - succeed_data_entry_count - 1
        substance_family_upload_log.end_time = datetime.now()
        substance_family_upload_log.status = 'success'
        substance_family_upload_log.save()
        # send substance_family_upload report to the user
        email_to = User.objects.filter(id=user_id).first().email
        subject = f'Substance family upload report'
        variables_dict = {
            'date': str(timezone.now().date()),
            'upload_report': f"{succeed_data_entry_count} substance families have been uploaded",
            'file_name': uploaded_file_name,
            'has_fail': has_fail,
            'failed_substance_family_list': failed_substance_family_list
        }
        template_id = settings.MAILJET_SUBSTANCE_FAMILY_UPLOAD_TEMPLATE_ID
        send_mail_via_mailjet_template(email_to, template_id, subject, variables_dict)
        print(f"substance_family_upload_task ..........success..........")
    except Exception as e:
        logger.error(str(e), exc_info=True)
        print(f"{uploaded_file_name} not found ....................")
        substance_family_upload_log.status = 'fail'
        substance_family_upload_log.end_time = datetime.now()
        substance_family_upload_log.save()
