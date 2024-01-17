from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema

import logging
from dateutil import parser
from django.utils import timezone
import openpyxl
from django.db.models import Q
from elasticsearch_dsl import Q as esQ
from rest_framework import views
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from rest_framework.response import Response
from rest_framework.filters import SearchFilter
from django_filters.rest_framework import DjangoFilterBackend

from rttregulation.filters import RegulationFilter, RegulatoryFrameworkFilter
from rttcore.permissions import is_staff_or_superuser
from rttorganization.models.models import Organization
from rttproduct.services.product_services import ProductServices
from rttregulation.models.core_models import Topic
from rttregulation.models.models import (Language, Status, Region, Url, RegulatoryFramework, MilestoneType,
                                         RegulationMilestone, IssuingBody, RegulationType, Regulation, Question,
                                         QuestionType, Answer, RegulationRatingLog, RegulatoryFrameworkRatingLog)
from rttregulation.serializers.serializers import (
    LanguageSerializer,
    StatusSerializer,
    RegionSerializer,
    UrlSerializer,
    RegulatoryFrameworkSerializer,
    MilestoneTypeSerializer, IssuingBodySerializer,
    RegulationTypeSerializer, RegulationSerializer,
    MilestoneSerializer, QuestionTypeSerializer, QuestionSerializer, AnswerSerializer,
    RegulatoryFrameworkDetailsSerializer, IssuingBodyDetailSerializer, RegulationDetailSerializer,
    ImpactAssessmentRegulationSerializer, TopicSerializer, RegulationRatingLogSerializer,
    RegulatoryFrameworkRatingLogSerializer, MilestoneListSerializer, RegulatoryFrameworkListSerializer,
    RegulationListSerializer, AnswerUpdateSerializer, AnswerPinSerializer)
from rttproduct.documents import ProductCategoryDocument, MaterialCategoryDocument
from rttregulation.documents import RegulationDocument, RegulatoryFrameworkDocument
from rttorganization.documents import OrganizationDocument

logger = logging.getLogger(__name__)


class LanguageViewSet(viewsets.ModelViewSet):
    queryset = Language.objects.all()
    serializer_class = LanguageSerializer
    permission_classes = [IsAdminUser]


class StatusViewSet(viewsets.ModelViewSet):
    queryset = Status.objects.all()
    serializer_class = StatusSerializer
    permission_classes = [IsAdminUser]


class RegionViewSet(viewsets.ModelViewSet):
    queryset = Region.objects.all()
    serializer_class = RegionSerializer
    permission_classes = [IsAdminUser]

    @action(detail=False, methods=['get'], url_path="options")
    def region_options(self, request):
        """
        doc: https://chemycal.atlassian.net/browse/RTT-1015
        """
        try:
            search_keyword = request.GET.get('search', None)
            response = []
            region_queryset = Region.objects.all()
            if search_keyword:
                region_queryset = region_queryset.filter(name__icontains=search_keyword)
            for region in region_queryset:
                response.append({
                    'id': region.id,
                    'name': region.name
                })
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class UrlViewSet(viewsets.ModelViewSet):
    queryset = Url.objects.all()
    serializer_class = UrlSerializer
    permission_classes = [IsAdminUser]
    lookup_field = 'id'


class RegulatoryFrameworkViewSet(viewsets.ModelViewSet):
    queryset = RegulatoryFramework.objects.all().select_related('language', 'status', 'issuing_body', ) \
        .prefetch_related('regions', 'documents', 'material_categories', 'product_categories', 'urls').distinct()
    serializer_classes = {
        'list': RegulatoryFrameworkListSerializer,
        'retrieve': RegulatoryFrameworkDetailsSerializer,
    }
    default_serializer_class = RegulatoryFrameworkSerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'
    filter_backends = [SearchFilter, DjangoFilterBackend]
    filter_class = RegulatoryFrameworkFilter
    search_fields = ['name', 'description']

    def get_serializer_class(self):
        return self.serializer_classes.get(self.action, self.default_serializer_class)

    @action(detail=True, methods=['get'])
    def get_relevant_organizations(self, request, id=None):
        regulatory_framework_id = int(id)
        try:
            regulation_id = []
            regulation_document_queryset = RegulationDocument.search().filter(
                esQ('match', review_status='o') &
                esQ('match', regulatory_framework__id=regulatory_framework_id)
            ).source(['id'])
            regulation_document_queryset = regulation_document_queryset[0:regulation_document_queryset.count()]
            for regulation in regulation_document_queryset:
                regulation_id.append(regulation.id)

            product_category_ids = []
            material_categories_ids = []
            product_category_document_queryset = ProductCategoryDocument.search().filter(
                esQ('nested',
                    path='product_cat_reg_framework',
                    query=esQ('match', product_cat_reg_framework__id=regulatory_framework_id)) |
                esQ('nested',
                    path='regulation_product_categories',
                    query=esQ('terms', regulation_product_categories__id=regulation_id))
            ).source(['id'])
            product_category_document_queryset = product_category_document_queryset[
                                                 0:product_category_document_queryset.count()]
            for product_category in product_category_document_queryset:
                product_category_ids.append(product_category.id)
            product_category_ids = ProductServices().get_all_tree_child_product_category_ids(product_category_ids)

            material_categories_document_queryset = MaterialCategoryDocument.search().filter(
                esQ('nested',
                    path='material_cat_reg_framework',
                    query=esQ('match', material_cat_reg_framework__id=regulatory_framework_id)) |
                esQ('nested',
                    path='regulation_material_categories',
                    query=esQ('terms', regulation_material_categories__id=regulation_id))
            ).source(['id'])
            material_categories_document_queryset = material_categories_document_queryset[
                                                    0:material_categories_document_queryset.count()]
            for material_categories in material_categories_document_queryset:
                material_categories_ids.append(material_categories.id)

            framework_topic_ids = []
            if not product_category_ids and not material_categories_ids:
                framework_topic_queryset = Topic.objects.filter(
                    regulatory_framework_topics=regulatory_framework_id).distinct()
                framework_topic_queryset = framework_topic_queryset[0:framework_topic_queryset.count()]
                for topic in framework_topic_queryset:
                    framework_topic_ids.append(topic.id)

            organization_document_queryset = OrganizationDocument.search().filter(
                esQ('nested',
                    path='product_organization.product_categories',
                    query=esQ('terms', product_organization__product_categories__id=product_category_ids)) |
                esQ('nested',
                    path='product_organization.material_categories',
                    query=esQ('terms', product_organization__material_categories__id=material_categories_ids)) |
                esQ('nested',
                    path='industries.topics',
                    query=esQ('terms', industries__topics__id=framework_topic_ids))

            ).source(['id', 'name'])
            organization_document_queryset = organization_document_queryset[0:organization_document_queryset.count()]

            if not organization_document_queryset:
                return Response(status=status.HTTP_204_NO_CONTENT)
            response_data = []
            for organization_document in organization_document_queryset:
                organization_data = {
                    'id': organization_document.id,
                    'name': organization_document.name,
                }
                response_data.append(organization_data)
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as ex:
            print(ex)
        return Response({'message': 'An error occurred'}, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'product_categories': openapi.Schema(type=openapi.TYPE_ARRAY, description='List of product_category ID',
                                                 items=openapi.Schema(type=openapi.TYPE_INTEGER)),
            'material_categories': openapi.Schema(type=openapi.TYPE_ARRAY, description='List of material_category ID',
                                                  items=openapi.Schema(type=openapi.TYPE_INTEGER)),
            'regulations': openapi.Schema(type=openapi.TYPE_ARRAY, description='List of regulations ID',
                                          items=openapi.Schema(type=openapi.TYPE_INTEGER)),
            'topics': openapi.Schema(type=openapi.TYPE_ARRAY, description='List of topics ID',
                                     items=openapi.Schema(type=openapi.TYPE_INTEGER))
        }
    ))
    @action(detail=False, methods=['post'], url_path="get_relevant_organizations")
    def relevant_organizations_framework(self, request):
        """
        doc: https://chemycal.atlassian.net/browse/RTT-672
        """
        try:
            filters = {
                'product_categories': request.data.get('product_categories', None),
                'material_categories': request.data.get('material_categories', None),
                'regulations': request.data.get('regulations', None)
            }
            topic_id_list = request.data.get('topics', None)
            product_category_ids = set()
            material_category_ids = set()

            regulation_doc_qs = RegulationDocument.search().filter('terms', id=filters['regulations'])
            regulation_doc_qs = regulation_doc_qs[0:regulation_doc_qs.count()]
            for regulation in regulation_doc_qs:
                for product_category in regulation.product_categories:
                    product_category_ids.add(product_category.id)
                for material_category in regulation.material_categories:
                    material_category_ids.add(material_category.id)
            product_category_ids.update(filters['product_categories'])
            product_category_ids = list(product_category_ids)
            product_category_ids = ProductServices().get_all_tree_child_product_category_ids(product_category_ids)
            material_category_ids.update(filters['material_categories'])
            material_category_ids = list(material_category_ids)

            if product_category_ids or material_category_ids:
                topic_id_list = []
            response = []
            organization_document_queryset = OrganizationDocument.search().filter(
                esQ('nested',
                    path='product_organization.product_categories',
                    query=esQ('terms', product_organization__product_categories__id=product_category_ids)) |
                esQ('nested',
                    path='product_organization.material_categories',
                    query=esQ('terms', product_organization__material_categories__id=material_category_ids)) |
                esQ('nested',
                    path='industries.topics',
                    query=esQ('terms', industries__topics__id=topic_id_list))
            ).source(['id', 'name'])
            organization_document_queryset = organization_document_queryset[0:organization_document_queryset.count()]
            for organization in organization_document_queryset:
                organization_obj = {
                    'id': organization.id,
                    'name': organization.name
                }
                response.append(organization_obj)
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path="dropdown-data")
    def dropdown_data(self, request):
        """
        doc: https://chemycal.atlassian.net/browse/RTT-751
        """
        try:
            search_keyword = request.GET.get('search', None)
            response = []
            framework_queryset = RegulatoryFramework.objects.prefetch_related('regions')
            if search_keyword:
                framework_queryset = framework_queryset.filter(name__icontains=search_keyword)
            framework_queryset = framework_queryset.filter(~Q(review_status='d'))
            for framework in framework_queryset:
                response.append({
                    'id': framework.id,
                    'name': framework.name,
                    'regions': [{'id': region.id, 'name': region.name} for region in framework.regions.all()]
                })
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path="options")
    def regulatory_framework_option(self, request):
        """
        doc: https://chemycal.atlassian.net/browse/RTT-1015
        """
        try:
            search_keyword = request.GET.get('search', None)
            response = []
            framework_queryset = RegulatoryFramework.objects.all()
            if search_keyword:
                framework_queryset = framework_queryset.filter(name__icontains=search_keyword)
            for framework in framework_queryset:
                response.append({
                    'id': framework.id,
                    'name': framework.name
                })
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class MilestoneTypeViewSet(viewsets.ModelViewSet):
    queryset = MilestoneType.objects.all()
    serializer_class = MilestoneTypeSerializer
    permission_classes = [IsAdminUser]
    lookup_field = 'id'


class MilestoneViewSet(viewsets.ModelViewSet):
    queryset = RegulationMilestone.objects.all()
    serializer_class = MilestoneSerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'

    def list(self, request, **kwargs):
        queryset = self.get_queryset()
        rf_id = self.request.GET.get('rf_id', None)
        r_id = self.request.GET.get('r_id', None)
        if rf_id is not None:
            queryset = self.queryset.filter(regulatory_framework=rf_id).order_by('-from_date')
        if r_id is not None:
            queryset = self.queryset.filter(regulation=r_id).order_by('-from_date')
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_paginated_response(MilestoneListSerializer(page, many=True).data)
        else:
            serializer = MilestoneListSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def create(self, request, *args, **kwargs):
        data = request.data
        serializer = MilestoneSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # Upload to the system a large amount of milestones
    # docs: https://chemycal.atlassian.net/browse/RTT-474
    @action(detail=False, methods=['post'])
    def bulk_regulatory_framework(self, request, **kwargs):
        file_request = request.FILES['file']
        if str(file_request.content_type) != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return_msg = {
                'message': 'Invalid File Type',
            }
            return Response(return_msg, status=status.HTTP_400_BAD_REQUEST)

        worksheet_data = openpyxl.load_workbook(file_request).worksheets[0]
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[1].value is None or row[4].value is None:
                continue
            try:
                created = RegulationMilestone.objects.create(
                    regulatory_framework_id=row[0].value,
                    type_id=row[1].value,
                    from_date=(parser.parse(str(row[2].value)) if row[2].value is not None else None),
                    to_date=(parser.parse(str(row[3].value)) if row[3].value is not None else None),
                    name=row[4].value,
                    description=(row[5].value if row[5].value is not None else '')
                )
            except Exception as e:
                print(e)

        return Response({'message': 'Bulk Data Created'}, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def bulk_regulation(self, request, **kwargs):
        file_request = request.FILES['file']
        if str(file_request.content_type) != "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            return_msg = {
                'message': 'Invalid File Type',
            }
            return Response(return_msg, status=status.HTTP_400_BAD_REQUEST)

        worksheet_data = openpyxl.load_workbook(file_request).worksheets[0]
        count = 0
        for row in worksheet_data.iter_rows():
            count += 1
            if count == 1 or row[0].value is None or row[1].value is None or row[4].value is None:
                continue
            try:
                created = RegulationMilestone.objects.create(
                    regulation_id=row[0].value,
                    type_id=row[1].value,
                    from_date=(parser.parse(str(row[2].value)) if row[2].value is not None else None),
                    to_date=(parser.parse(str(row[3].value)) if row[3].value is not None else None),
                    name=row[4].value,
                    description=(row[5].value if row[5].value is not None else '')
                )
            except Exception as e:
                print(e)
        return Response({'message': 'Bulk Data Created'}, status=status.HTTP_201_CREATED)


class IssuingBodyViewSet(viewsets.ModelViewSet):
    queryset = IssuingBody.objects.all().select_related('region')
    serializer_classes = {
        'list': IssuingBodyDetailSerializer,
        'retrieve': IssuingBodyDetailSerializer,
    }
    default_serializer_class = IssuingBodySerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'

    filter_backends = [SearchFilter, ]
    search_fields = ['name', 'description']

    def get_serializer_class(self):
        return self.serializer_classes.get(self.action, self.default_serializer_class)


class RegulationTypeViewSet(viewsets.ModelViewSet):
    queryset = RegulationType.objects.all()
    serializer_class = RegulationTypeSerializer
    permission_classes = [IsAdminUser]


class RegulationViewSet(viewsets.ModelViewSet):
    queryset = Regulation.objects.all().select_related('type', 'status')
    serializer_classes = {
        'list': RegulationListSerializer,
        'retrieve': RegulationDetailSerializer,
    }
    default_serializer_class = RegulationSerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'
    filter_backends = [SearchFilter, DjangoFilterBackend]
    filter_class = RegulationFilter
    search_fields = ['name', 'description']

    def get_serializer_class(self):
        return self.serializer_classes.get(self.action, self.default_serializer_class)

    @action(detail=True, methods=['get'])
    def details(self, request, **kwargs):
        regulation = self.get_object()
        serializer = RegulationSerializer(regulation)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def get_relevant_organizations(self, request, id=None):
        regulation_id = int(id)
        try:
            product_category_ids = []
            material_categories_ids = []
            product_category_document_queryset = ProductCategoryDocument.search().filter(
                esQ('nested',
                    path='regulation_product_categories',
                    query=esQ('match', regulation_product_categories__id=regulation_id))
            ).source(['id'])
            product_category_document_queryset = product_category_document_queryset[
                                                 0:product_category_document_queryset.count()]
            for product_category in product_category_document_queryset:
                product_category_ids.append(product_category.id)
            product_category_ids = ProductServices().get_all_tree_child_product_category_ids(product_category_ids)

            material_categories_document_queryset = MaterialCategoryDocument.search().filter(
                esQ('nested',
                    path='regulation_material_categories',
                    query=esQ('match', regulation_material_categories__id=regulation_id))
            ).source(['id'])
            material_categories_document_queryset = material_categories_document_queryset[
                                                    0:material_categories_document_queryset.count()]
            for material_categories in material_categories_document_queryset:
                material_categories_ids.append(material_categories.id)

            regulation_topic_ids = []
            if not product_category_ids and not material_categories_ids:
                regulation_topic_queryset = Topic.objects.filter(regulation_topics=regulation_id).distinct()
                regulation_topic_queryset = regulation_topic_queryset[0:regulation_topic_queryset.count()]
                for topic in regulation_topic_queryset:
                    regulation_topic_ids.append(topic.id)

            organization_document_queryset = OrganizationDocument.search().filter(
                esQ('nested',
                    path='product_organization.product_categories',
                    query=esQ('terms', product_organization__product_categories__id=product_category_ids)) |
                esQ('nested',
                    path='product_organization.material_categories',
                    query=esQ('terms', product_organization__material_categories__id=material_categories_ids)) |
                esQ('nested',
                    path='industries.topics',
                    query=esQ('terms', industries__topics__id=regulation_topic_ids))
            ).source(['id', 'name'])
            organization_document_queryset = organization_document_queryset[0:organization_document_queryset.count()]

            if not organization_document_queryset:
                return Response(status=status.HTTP_204_NO_CONTENT)
            response_data = []
            for organization_document in organization_document_queryset:
                organization_data = {
                    'id': organization_document.id,
                    'name': organization_document.name,
                }
                response_data.append(organization_data)
            return Response(response_data, status=status.HTTP_200_OK)

        except Exception as ex:
            print(ex)
        return Response({'message': 'An error occurred'}, status=status.HTTP_400_BAD_REQUEST)

    @swagger_auto_schema(request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'product_categories': openapi.Schema(type=openapi.TYPE_ARRAY, description='List of product_category ID',
                                                 items=openapi.Schema(type=openapi.TYPE_INTEGER)),
            'material_categories': openapi.Schema(type=openapi.TYPE_ARRAY, description='List of material_category ID',
                                                  items=openapi.Schema(type=openapi.TYPE_INTEGER)),
            'topics': openapi.Schema(type=openapi.TYPE_ARRAY, description='List of topics ID',
                                     items=openapi.Schema(type=openapi.TYPE_INTEGER))
        }
    ))
    @action(detail=False, methods=['post'], url_path="get_relevant_organizations")
    def relevant_organizations_regulation(self, request):
        """
        https://chemycal.atlassian.net/browse/RTT-673
        """
        try:
            filters = {
                'product_categories': request.data.get('product_categories', None),
                'material_categories': request.data.get('material_categories', None)
            }
            topic_id_list = request.data.get('topics', None)
            response = []
            product_category_ids = filters['product_categories']
            product_category_ids = ProductServices().get_all_tree_child_product_category_ids(product_category_ids)
            if product_category_ids or request.data.get('material_categories', None):
                topic_id_list = []
            organization_document_queryset = OrganizationDocument.search().filter(
                esQ('nested',
                    path='product_organization.product_categories',
                    query=esQ('terms', product_organization__product_categories__id=product_category_ids)) |
                esQ('nested',
                    path='product_organization.material_categories',
                    query=esQ('terms', product_organization__material_categories__id=filters['material_categories'])) |
                esQ('nested',
                    path='industries.topics',
                    query=esQ('terms', industries__topics__id=topic_id_list))
            ).source(['id', 'name'])
            organization_document_queryset = organization_document_queryset[0:organization_document_queryset.count()]
            for organization in organization_document_queryset:
                organization_obj = {
                    'id': organization.id,
                    'name': organization.name
                }
                response.append(organization_obj)
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path="dropdown-data")
    def dropdown_data(self, request):
        """
        doc: https://chemycal.atlassian.net/browse/RTT-751
        """
        try:
            search_keyword = request.GET.get('search', None)
            response = []
            regulation_queryset = Regulation.objects.select_related('regulatory_framework')
            if search_keyword:
                regulation_queryset = regulation_queryset.filter(name__icontains=search_keyword)
            regulation_queryset = regulation_queryset.filter(~Q(review_status='d'))
            for regulation in regulation_queryset:
                region_list = []
                if regulation.regulatory_framework:
                    for region in regulation.regulatory_framework.regions.all():
                        region_list.append({'id': region.id, 'name': region.name})
                response.append({
                    'id': regulation.id,
                    'name': regulation.name,
                    'regions': region_list
                })
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path="options")
    def regulation_option(self, request):
        """
        doc: https://chemycal.atlassian.net/browse/RTT-1015
        """
        try:
            search_keyword = request.GET.get('search', None)
            response = []
            regulation_queryset = Regulation.objects.all()
            if search_keyword:
                regulation_queryset = regulation_queryset.filter(name__icontains=search_keyword)
            for regulation in regulation_queryset:
                response.append({
                    'id': regulation.id,
                    'name': regulation.name
                })
            return Response(response, status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class QuestionTypeViewSet(viewsets.ModelViewSet):
    queryset = QuestionType.objects.all()
    serializer_class = QuestionTypeSerializer
    permission_classes = [IsAdminUser]


class ImpactAssessmentQuestion(viewsets.ModelViewSet):
    queryset = Question.objects.all()
    serializer_class = QuestionSerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'

    def get_queryset(self):
        if is_staff_or_superuser(self.request):
            organization_id = self.request.GET.get('organization_id', None)
            if organization_id:
                return self.queryset.filter(organization_id=organization_id)
            else:
                return self.queryset
        else:
            organization_id = self.request.user.organization_id
            return self.queryset.filter(organization_id=organization_id)


class ImpactAssessmentAnswer(viewsets.ModelViewSet):
    queryset = Answer.objects.all()
    serializer_class = AnswerSerializer
    serializer_classes = {
        'retrieve': AnswerSerializer,
        'update': AnswerUpdateSerializer,
        'partial_update': AnswerUpdateSerializer,
    }
    default_serializer_class = AnswerSerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'

    def get_serializer_class(self):
        return self.serializer_classes.get(self.action, self.default_serializer_class)

    def get_queryset(self):
        framework_id = self.request.GET.get('framework_id', None)
        regulation_id = self.request.GET.get('regulation_id', None)
        if framework_id:
            return self.queryset.filter(regulatory_framework_id=framework_id)
        if regulation_id:
            return self.queryset.filter(regulation_id=regulation_id)
        return self.queryset

    def list(self, request, *args, **kwargs):
        questions = Question.objects.filter(organization_id=request.user.organization_id)
        response = {
            'count': 0,
            'results': []
        }
        for question in questions:
            answer = self.get_queryset().filter(question_id=question.id).order_by('-id').first()
            response['results'].append(self.get_serializer(answer).data)
        response['count'] = response['results'].__len__()
        return Response(response, status=status.HTTP_200_OK)

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        data['answered_by'] = request.user.id
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        user_id = request.user.id
        '''
        Before setting new pin_by, have to remove old pin_by for that question. 
        Cause only one pined answer may have in question. 
        https://chemycal.atlassian.net/browse/RTT-883
        '''
        if request.data.get('pin_by', None):
            serializer = AnswerPinSerializer(instance, data=request.data, partial=True)
            if serializer.is_valid(raise_exception=True):
                try:
                    answer_qs = Answer.objects.filter(question_id=instance.question)
                    if instance.regulation:
                        answer_qs.filter(regulation_id=instance.regulation).update(pin_by=None)
                    else:
                        answer_qs.filter(regulatory_framework_id=instance.regulatory_framework).update(pin_by=None)
                except Exception as ex:
                    logger.error(str(ex), exc_info=True)
                self.perform_update(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        if user_id != instance.answered_by.id:
            return Response({"message": "Only answer owner has the permission to edit"},
                            status=status.HTTP_403_FORBIDDEN)
        else:
            if request.data.get('edited', None):
                request.data['edited'] = instance.edited if instance.edited else None
            if request.data.get('answer_text', None):
                if instance.answer_text != request.data.get('answer_text'):
                    request.data['edited'] = timezone.now()
            serializer = self.get_serializer(instance, data=request.data, partial=True)
            if serializer.is_valid(raise_exception=True):
                self.perform_update(serializer)
                return Response(serializer.data, status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def destroy(self, request, *args, **kwargs):
        try:
            instance = self.get_object()
            if instance.answered_by.id != request.user.id:
                return Response({'message': 'you do not have permission to delete!'}, status=status.HTTP_403_FORBIDDEN)
            self.perform_destroy(instance)
            return Response(status=status.HTTP_200_OK)
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "server error"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


class RegulatoryFrameworkUrlApiView(views.APIView):
    permission_classes = (IsAuthenticated,)

    @staticmethod
    def post(request):
        data = request.data
        url_data = {'text': data['text'], 'description': data['description']}
        url_serializer = UrlSerializer(data=url_data)
        if url_serializer.is_valid():
            url = url_serializer.save()
            reg_framework = RegulatoryFramework.objects.filter(id=data['id']).first()
            reg_framework.urls.add(url)
            return Response(url_serializer.data, status=status.HTTP_201_CREATED)
        return Response(status=status.HTTP_400_BAD_REQUEST)


class GetImpactAssessment:

    def impact_assessment(self, pk):
        organizations = list(Organization.objects.prefetch_related(
            'product_organization__material_categories__regulation_material_categories',
            'product_organization__product_categories__regulation_product_categories'
        ).filter(
            Q(product_organization__product_categories__regulation_product_categories__id=pk) |
            Q(product_organization__material_categories__regulation_material_categories__id=pk)
        ).values('id', 'name').distinct())

        results = {'regulation': {}, 'questions': []}

        regulation = Regulation.objects.select_related('regulatory_framework__issuing_body').prefetch_related(
            'regulatory_framework__regions').filter(id=pk).first()
        regulation_serializer = ImpactAssessmentRegulationSerializer(regulation)
        issuing_body = regulation_serializer.data['regulatory_framework'].get('issuing_body')
        issuing_body_name = None if issuing_body is None else issuing_body.get('name')
        results['regulation'] = {
            'regulation_name': regulation_serializer.data.get('name'),
            'issuing_body': issuing_body_name,
            'regions': regulation_serializer.data['regulatory_framework'].get('regions'),
        }

        return self.get_result(results, organizations)

    def regulatory_framework_impact_assessment(self, rf_id):
        results = {'questions': []}

        organizations = list(Organization.objects.prefetch_related(
            'product_organization__material_categories__regulation_material_categories',
            'product_organization__product_categories__regulation_product_categories'
        ).filter(
            Q(product_organization__product_categories__product_cat_reg_framework__id=rf_id) |
            Q(product_organization__material_categories__material_cat_reg_framework__id=rf_id)
        ).values('id', 'name').distinct())

        return self.get_result(results, organizations)

    @staticmethod
    def get_result(results, organizations):
        for organization in organizations:
            questions = Question.objects.filter(Q(organization=organization['id'])).select_related('organization')
            questions_with_ans = questions.filter(Q(answer__isnull=False))
            questions_serializer = QuestionSerializer(questions, many=True)
            results['questions'].append({
                'organization_name': organization['name'],
                'questions_count': questions.count(),
                'answered': questions_with_ans.count(),
                'questions': questions_serializer.data
            })
        return results


class ImpactAssessmentApiView(views.APIView):
    permission_classes = (IsAuthenticated,)

    @staticmethod
    def get(request):
        regulation_id = request.GET.get('r_id', None)
        regulatory_framework_id = request.GET.get('rf_id', None)
        obj = GetImpactAssessment()
        if regulation_id:
            results = obj.impact_assessment(regulation_id)
        else:
            results = obj.regulatory_framework_impact_assessment(regulatory_framework_id)
        return Response({'data': results}, status=status.HTTP_201_CREATED)


class ImpactAssessmentListApiView(views.APIView):
    permission_classes = (IsAuthenticated,)

    @staticmethod
    def get(request):
        obj = GetImpactAssessment()
        regulations = Regulation.objects.all().values_list('id', flat=True)
        count = regulations.count()
        lower_limit = int(request.GET.get('pageSize', 0)) * (int(request.GET.get('pageNumber', 1)) - 1)
        upper_limit = int(request.GET.get('pageSize', 0)) * (int(request.GET.get('pageNumber', 1)))
        results = []
        if upper_limit != 0:
            for regulation in regulations[lower_limit:upper_limit]:
                results.append(obj.impact_assessment(regulation))
        else:
            for regulation in regulations:
                results.append(obj.impact_assessment(regulation))
        return Response({'results': results, 'count': count}, status=status.HTTP_201_CREATED)


class TopicViewSet(viewsets.ModelViewSet):
    queryset = Topic.objects.all()
    serializer_class = TopicSerializer
    permission_classes = [IsAdminUser]
    lookup_field = 'id'


class RegulationRatingViewSet(viewsets.ModelViewSet):
    queryset = RegulationRatingLog.objects.all()
    serializer_class = RegulationRatingLogSerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        if request.data.get('regulation', None):
            regulation_id = request.data['regulation']
            regulation_rating_log = RegulationRatingLog.objects \
                .filter(regulation_id=regulation_id, organization_id=request.data.get('organization')) \
                .order_by('-created').first()
            if regulation_rating_log:
                if request.data.get('rating', None) and request.data['rating'] == regulation_rating_log.rating:
                    return Response({"message": "previous and current rating is same"},
                                    status=status.HTTP_400_BAD_REQUEST)
                data['prev_rating'] = regulation_rating_log.rating
                data['parent'] = regulation_rating_log.id
            else:
                data['prev_rating'] = None
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class RegulatoryFrameworkRatingViewSet(viewsets.ModelViewSet):
    queryset = RegulatoryFrameworkRatingLog.objects.all()
    serializer_class = RegulatoryFrameworkRatingLogSerializer
    permission_classes = (IsAuthenticated,)
    lookup_field = 'id'

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        if request.data.get('regulatory_framework', None):
            framework_id = request.data['regulatory_framework']
            regulatory_framework_rating_log = RegulatoryFrameworkRatingLog.objects\
                .filter(regulatory_framework_id=framework_id, organization_id=request.data.get('organization'))\
                .order_by('-created').first()
            if regulatory_framework_rating_log:
                if request.data.get('rating', None) and request.data['rating'] == regulatory_framework_rating_log.rating:
                    return Response({"message": "previous and current rating is same"},
                                    status=status.HTTP_400_BAD_REQUEST)
                data['prev_rating'] = regulatory_framework_rating_log.rating
                data['parent'] = regulatory_framework_rating_log.id
            else:
                data['prev_rating'] = None
        serializer = self.get_serializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class ImpactAssessmentAnswerListCreateView(views.APIView):
    permission_classes = (IsAuthenticated,)

    @staticmethod
    def post(request):
        data_list = request.data
        serializers = AnswerSerializer(data=data_list, many=True)
        if serializers.is_valid():
            serializers.save()
            return Response(serializers.data, status=status.HTTP_201_CREATED)
        else:
            return Response(serializers.errors, status=status.HTTP_400_BAD_REQUEST)
