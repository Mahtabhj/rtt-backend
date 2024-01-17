import logging

from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response

from rttregulation.services.what_next_milestone_service import WhatNextMilestoneService
from rttregulation.services.regulation_tagged_region_service import RegulationTaggedRegionService

logger = logging.getLogger(__name__)


class ExportMilestoneDataAPIView(APIView):
    permission_classes = (IsAuthenticated, )

    @staticmethod
    def post(request):
        """
        doc: https://chemycal.atlassian.net/browse/RTT-1334
        """
        try:
            period = str(request.data.get('period', ''))
            filters = {
                'period_start': period + '-01-01' if period else None,
                'period_end': period + '-12-31' if period else None,
                'regulatory_frameworks': request.data.get('regulatory_frameworks', None),
                'regulations': request.data.get('regulations', None),
                'milestone_types': request.data.get('milestone_types', None),
                'regions': request.data.get('regions', None),
                'product_categories': request.data.get('product_categories', None),
                'material_categories': request.data.get('material_categories', None),
                'related_products': request.data.get('related_products', None),
                'status': request.data.get('status', None),
                'topics': request.data.get('topics', None)
            }
            organization_id = request.user.organization_id
            search_keyword = request.data.get('search', None)
            milestone_doc_queryset = WhatNextMilestoneService().get_what_next_filtered_milestone_document_queryset(
                organization_id, filters, search_keyword)
            milestone_doc_queryset = milestone_doc_queryset[0:milestone_doc_queryset.count()]

            workbook = Workbook()
            work_sheet = workbook.active
            work_sheet.title = 'Milestone'
            work_sheet.cell(column=7, row=1, value='Blank if framework')
            work_sheet.cell(column=8, row=1, value='Blank if regulation')
            header_name_list = ['Milestone ID', 'Milestone name', 'Milestone description', 'Milestone type',
                                'Regions', 'Date', 'Regulation Name', 'Framework Name']
            for idx, name in enumerate(header_name_list):
                work_sheet.cell(column=idx + 1, row=2, value=name)
            row_idx = 3
            for milestone in milestone_doc_queryset:
                work_sheet.cell(column=1, row=row_idx, value=milestone.id)
                work_sheet.cell(column=2, row=row_idx, value=milestone.name)
                work_sheet.cell(column=3, row=row_idx, value=milestone.description if milestone.description else '')
                if milestone.type:
                    work_sheet.cell(column=4, row=row_idx, value=milestone.type.name)
                work_sheet.cell(column=6, row=row_idx, value=str(
                    milestone.from_date.date())if milestone.from_date else None)
                regions = ""
                if milestone.regulation:
                    work_sheet.cell(column=7, row=row_idx, value=milestone.regulation.name)
                    if milestone.regulation.regulatory_framework:
                        region_list = RegulationTaggedRegionService().get_region_data(
                            milestone.regulation.regulatory_framework.id)
                        for region in region_list:
                            regions += region['name'] if len(regions) == 0 else (', '+region['name'])
                if milestone.regulatory_framework:
                    work_sheet.cell(column=8, row=row_idx, value=milestone.regulatory_framework.name)
                    for region in milestone.regulatory_framework.regions:
                        regions += region.name if len(regions) == 0 else (', '+region.name)
                work_sheet.cell(column=5, row=row_idx, value=regions)

                row_idx += 1

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="export_milestone.xlsx"'
            workbook.save(response)
            return response
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "INTERNAL_SERVER_ERROR!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
