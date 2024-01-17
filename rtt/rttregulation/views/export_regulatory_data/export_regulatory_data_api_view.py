import logging

from django.http import HttpResponse
from openpyxl import Workbook
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework.views import APIView
from rest_framework.response import Response

from rttcore.permissions import IsSuperUserOrStaff
from rttregulation.documents import RegulatoryFrameworkDocument, RegulationDocument
from rttlimitManagement.models import Exemption
from rttproduct.models.core_models import Industry

logger = logging.getLogger(__name__)
workbook = Workbook()


class ExportRegulatoryDataAPIView(APIView):
    permission_classes = (IsAuthenticated, IsSuperUserOrStaff, )

    def post(self, request):
        try:
            work_sheet = workbook.active
            work_sheet.title = 'regulation, framework database'
            work_sheet.cell(column=1, row=1, value='related to current framework or framework of the regulation')
            work_sheet.cell(column=3, row=1, value='blank if it is a framework')
            work_sheet.cell(column=5, row=1, value='Blank if framework')
            work_sheet.cell(column=6, row=1, value='data of framework or regulation')
            work_sheet.cell(column=9, row=1, value='blank if regulation')
            work_sheet.cell(column=10, row=1, value='From the framework if it is a regulation')
            work_sheet.cell(column=14, row=1, value='format: {material category} ({industry})')
            header_name_list = ['Framework ID', 'Framework name', 'Regulation ID ', 'Regulation name', 'Type',
                                'Description', 'Review status', 'Status', 'Issuing body', 'Regions', 'Topics',
                                'Description', 'Product categories', 'Material categories', 'has links',
                                'has documents', 'has milestones', 'has substances', 'has milestones with subs',
                                'has limits', 'has exemptions']
            for idx, name in enumerate(header_name_list):
                work_sheet.cell(column=idx + 1, row=2, value=name)
            row_idx = 3
            framework_doc_qs: RegulatoryFrameworkDocument = RegulatoryFrameworkDocument.search()
            framework_doc_qs = framework_doc_qs[0:framework_doc_qs.count()]
            for framework in framework_doc_qs:
                work_sheet.cell(column=1, row=row_idx, value=framework.id)
                work_sheet.cell(column=2, row=row_idx, value=framework.name)
                work_sheet.cell(column=6, row=row_idx, value=framework.description)
                work_sheet.cell(column=7, row=row_idx, value=framework.review_status)
                work_sheet.cell(column=8, row=row_idx, value=framework.status.name)
                work_sheet.cell(column=9, row=row_idx, value=framework.issuing_body.name)
                work_sheet.cell(column=10, row=row_idx, value=self.get_region_data(framework))
                work_sheet.cell(column=11, row=row_idx, value=self.get_topic_data(framework))
                work_sheet.cell(column=12, row=row_idx, value=framework.description)
                work_sheet.cell(column=13, row=row_idx, value=self.get_product_category_data(framework))
                work_sheet.cell(column=14, row=row_idx, value=self.get_material_category_data(framework))
                work_sheet.cell(column=15, row=row_idx, value=True if len(framework.urls) > 0 else False)
                work_sheet.cell(column=16, row=row_idx, value=True if len(framework.documents) > 0 else False)
                work_sheet.cell(column=17, row=row_idx,
                                value=True if len(framework.regulatory_framework_milestone) > 0 else False)
                work_sheet.cell(column=18, row=row_idx, value=True if len(framework.substances) > 0 else False)
                work_sheet.cell(column=19, row=row_idx, value=self.has_milestone_with_substance(framework, is_regulation=False))
                work_sheet.cell(column=20, row=row_idx, value=True if len(
                    framework.regulatory_framework_regulation_substance_limit) > 0 else False)
                work_sheet.cell(column=21, row=row_idx, value=self.has_exemption(framework.id, is_regulation=False))
                row_idx += 1

            regulation_doc_qs: RegulationDocument = RegulationDocument.search()
            regulation_doc_qs = regulation_doc_qs[0:regulation_doc_qs.count()]
            for regulation in regulation_doc_qs:
                work_sheet.cell(column=1, row=row_idx, value=regulation.regulatory_framework.id)
                work_sheet.cell(column=2, row=row_idx, value=regulation.regulatory_framework.name)
                work_sheet.cell(column=3, row=row_idx, value=regulation.id)
                work_sheet.cell(column=4, row=row_idx, value=regulation.name)
                work_sheet.cell(column=5, row=row_idx, value=regulation.type.name)
                work_sheet.cell(column=6, row=row_idx, value=regulation.description)
                work_sheet.cell(column=7, row=row_idx, value=regulation.review_status)
                work_sheet.cell(column=8, row=row_idx, value=regulation.status.name)
                work_sheet.cell(column=10, row=row_idx, value=self.get_region_data(regulation.regulatory_framework))
                work_sheet.cell(column=11, row=row_idx, value=self.get_topic_data(regulation))
                work_sheet.cell(column=12, row=row_idx, value=regulation.description)
                work_sheet.cell(column=13, row=row_idx, value=self.get_product_category_data(regulation))
                work_sheet.cell(column=14, row=row_idx, value=self.get_material_category_data(regulation))
                work_sheet.cell(column=15, row=row_idx, value=True if len(regulation.urls) > 0 else False)
                work_sheet.cell(column=16, row=row_idx, value=True if len(regulation.documents) > 0 else False)
                work_sheet.cell(column=17, row=row_idx,
                                value=True if len(regulation.regulation_milestone) > 0 else False)
                work_sheet.cell(column=18, row=row_idx, value=True if len(regulation.substances) > 0 else False)
                work_sheet.cell(column=19, row=row_idx, value=self.has_milestone_with_substance(regulation, is_regulation=True))
                work_sheet.cell(column=20, row=row_idx, value=True if len(
                    regulation.regulation_regulation_substance_limit) > 0 else False)
                work_sheet.cell(column=21, row=row_idx, value=self.has_exemption(regulation.id, is_regulation=True))
                row_idx += 1
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="export_regulatory_database.xlsx"'
            workbook.save(response)
            return response
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "Serve error!"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @staticmethod
    def get_region_data(framework_queryset):
        result = ''
        for region in framework_queryset.regions:
            if result:
                result += ', '
            result += region.name
        return result

    @staticmethod
    def get_topic_data(regulation_queryset):
        result = ''
        for topic in regulation_queryset.topics:
            if result:
                result += ', '
            result += topic.name
        return result

    @staticmethod
    def get_product_category_data(regulation_queryset):
        result = ''
        for product_category in regulation_queryset.product_categories:
            if result:
                result += ', '
            result += product_category.name
        return result

    @staticmethod
    def has_milestone_with_substance(regulation_queryset, is_regulation):
        if is_regulation:
            for milestone in regulation_queryset.regulation_milestone:
                if len(milestone.substances) > 0:
                    return True
        else:
            for milestone in regulation_queryset.regulatory_framework_milestone:
                if len(milestone.substances) > 0:
                    return True
        return False

    @staticmethod
    def get_material_category_data(regulation_queryset):
        result = ''
        for material_category in regulation_queryset.material_categories:
            if result:
                result += ', '

            industry_qs = Industry.objects.filter(material_category_industry__id=material_category.id).first()
            mat_cat_and_industry = material_category.name + ' (' + industry_qs.name + ')'
            result += mat_cat_and_industry
        return result

    @staticmethod
    def has_exemption(regulation_id, is_regulation):
        if is_regulation:
            return Exemption.objects.filter(regulation_id=regulation_id).exists()
        else:
            return Exemption.objects.filter(regulatory_framework_id=regulation_id).exists()
