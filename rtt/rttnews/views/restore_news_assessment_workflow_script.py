from drf_yasg import openapi
from drf_yasg.utils import swagger_auto_schema
import logging

from rest_framework.response import Response
from django.http import HttpResponse
from rest_framework import status
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook

from rttcore.permissions import IsSuperUserOrStaff
from rttcore.services.system_filter_service import SystemFilterService
from rttnews.models.models import NewsAssessmentWorkflow

logger = logging.getLogger(__name__)


class RestoreNewsAssessmentWorkflowScript(APIView):
    permission_classes = (IsAuthenticated, IsSuperUserOrStaff,)

    @swagger_auto_schema(request_body=openapi.Schema(
        type=openapi.TYPE_OBJECT,
        properties={
            'organization_id': openapi.Schema(type=openapi.TYPE_INTEGER, description='enter organization_id'),
        }
    ))
    def post(self, request, *args, **kwargs):
        try:
            if not request.data.get("organization_id", None):
                return Response({"message": "organization_id must be sent"}, status=status.HTTP_400_BAD_REQUEST)
            organization_id = request.data["organization_id"]
            system_filter_service = SystemFilterService()
            news_doc_qs = system_filter_service.get_system_filtered_news_document_queryset(organization_id)
            news_doc_qs = news_doc_qs[0:news_doc_qs.count()]
            workbook = Workbook()
            work_sheet = workbook.active
            work_sheet.title = 'Restore-news-assessment-workflow'
            work_sheet.cell(column=1, row=1, value='News ID')
            work_sheet.cell(column=2, row=1, value='Title')
            row_idx = 2
            for news in news_doc_qs:
                if not self.has_news_relevance(news.news_relevance, organization_id):
                    workflow_qs, created = NewsAssessmentWorkflow.objects.get_or_create(news_id=news.id, organization_id=organization_id)
                    if created:
                        work_sheet.cell(column=1, row=row_idx, value=news.id)
                        work_sheet.cell(column=2, row=row_idx, value=news.title)
                        row_idx += 1
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename="Restore-news-assessment-workflow.xlsx"'
            workbook.save(response)
            return response
        except Exception as ex:
            logger.error(str(ex), exc_info=True)
        return Response({"message": "INTERNAL_SERVER_ERROR"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @staticmethod
    def has_news_relevance(news_relevance_qs, organization_id):
        for news_relevance in news_relevance_qs:
            if news_relevance.organization.id == organization_id:
                return True
        return False
